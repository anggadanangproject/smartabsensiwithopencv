import cv2, os, numpy as np
import tkinter as tk
from tkinter import ttk
from tkinter import * 
import csv
from PIL import ImageTk, Image
from datetime import datetime
from tkinter.messagebox import askokcancel, showinfo, WARNING
import absensi as ab
import pegawai as pg
import time
from openpyxl import Workbook
from openpyxl import load_workbook

# global variable
global bulan, hasil
now = datetime.now()
bulan = now.strftime('%m-%Y')
hari = now.strftime('%d-%m-%Y')
wb = Workbook()
ws = wb.active

def absenMasuk():
    ab.absensi("Masuk")

def absenKeluar():
    ab.absensi("Keluar")

def laporan():
    ab.data()

def pegawai():
    pg.data()

def daftar():
    bersih()
    canvas.create_window(300,150, window=label1)
    canvas.create_window(300, 180, height=35, width=300, window=entry1)
    canvas.create_window(300, 240, window=btn4)
    canvas.create_window(550, 480, window=btn5)
    canvas.create_window(300, 440, window=(btn6))
    canvas.create_window(300, 400, window=(btn7))

def TambahNama(nama_pegawai):
    entry1.delete(0, END)
    data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
    sheet = data.active
    baris = 0

    for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
        baris +=1
        for row in range(0,len(value)):
            ws.cell(row=baris, column=row+1).value= value[row]
        print (baris)
    baris = baris+1
    ws.cell(row=baris, column=1).value= baris-1
    ws.cell(row=baris, column=2).value= nama_pegawai
    ws.cell(row=baris, column=3).value= "Pegawai"

    wb.save("data_pegawai/data_pegawai.xlsx")

def rekamDataWajah():
    wajahDir = 'datawajah'
    cam = cv2.VideoCapture(1)
    cam.set(3, 640)
    cam.set(4, 480) 
    id_pegawai = 0
    imagePaths = [os.path.join(wajahDir, f) for f in os.listdir(wajahDir)]
    for imagePath in imagePaths:
        id_pegawai = int(os.path.split(imagePath)[-1].split(".")[1])

    faceDetector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    eyeDetector = cv2.CascadeClassifier('haarcascade_eye.xml')

    face_id = id_pegawai+1
    nama_pegawai = entry1.get()
    ambilData = 0
    while (True):
        ret, img = cam.read()
        img = cv2.flip(img, 1)  # flip video image vertically
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = faceDetector.detectMultiScale(gray, 1.3, 5)
        k = cv2.waitKey(10) & 0xff  # Press 'ESC' for exiting video
        cv2.putText(img, str(ambilData)+"/100", (70,40), cv2.FONT_HERSHEY_PLAIN, 2,(0,255,0),2)
        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
            ambilData += 1
            # Save the captured image into the datasets folder
            cv2.imwrite("datawajah/user"+ '.'+str(face_id)+"." + str(ambilData) + ".jpg", gray[y:y + h, x:x + w])
                
        cv2.imshow("Rekam data wajah", img)
        if k == 27:
            break
        elif ambilData >= 100:
            face_id += 1
            ambilData = 0
            nama_pegawai = entry1.get()
            TambahNama(nama_pegawai)
            trainingWajah()  # Take 30 face sample and stop video
            break
    #selesai1()
    cam.release()
    cv2.destroyAllWindows()  # untuk menghapus data yang sudah dibaca


def trainingWajah():
    answer = askokcancel(
            title='Konfirmasi',
            message='Apakah anda ingin melakukan training data?',
            icon=WARNING)

    if answer:
        # Path for face image database
        path = 'datawajah'
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")

        # function to get the images and label data
        def getImagesAndLabels(path):

            imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
            faceSamples = []
            ids = []
            loop=0
            id=1

            for imagePath in imagePaths:
                loop+=1

                PIL_img = Image.open(imagePath).convert('L')  # convert it to grayscale
                img_numpy = np.array(PIL_img, 'uint8')

                print(str(id)+""+str(loop))
                faces = detector.detectMultiScale(img_numpy)

                for (x, y, w, h) in faces:
                    faceSamples.append(img_numpy[y:y + h, x:x + w])
                    ids.append(id)

                if loop==100:
                    id = id+1

            return faceSamples, ids

        #print("\n [INFO] Training faces. It will take a few seconds. Wait ...")
        faces, ids = getImagesAndLabels(path)
        recognizer.train(faces, np.array(ids))

        # Save the model into trainer/trainer.yml
        recognizer.write('latihwajah/trainer.yml')  # recognizer.save() worked on Mac, but not on Pi
        showinfo(
                title='Informasi',
                message='Training data sudah berhasil!')
        # Print the numer of faces trained and end program
        #print("\n [INFO] {0} faces trained. Exiting Program".format(len(np.unique(ids))))
        #selesai2()

def ambilGambar():
    if entry1.get() !="":
        answer = askokcancel(
            title='Konfirmasi',
            message='Apakah anda yakin menambahkan data ini?',
            icon=WARNING)

        if answer:
            rekamDataWajah();
    else:
        showinfo(
            title='Peringatan',
            message='Tuliskan nama pegawai terlebih dahulu!')

def home():
    global btn, btn1, btn2, btn3
    bersih()
    canvas.create_window(190, 200, window=(btn))
    canvas.create_window(410, 200, window=(btn1))
    canvas.create_window(190, 330, window=(btn2))
    canvas.create_window(410, 330, window=(btn3))

def bersih():
    canvas.create_window(-100, -100, window=(btn))
    canvas.create_window(-100, -100, window=(btn1))
    canvas.create_window(-100, -100, window=(btn2))
    canvas.create_window(-100, -100, window=(btn3))
    #pendaftaran
    canvas.create_window(-100, -100, window=(label1))
    canvas.create_window(-100, -100, window=(entry1))
    canvas.create_window(-100, -100, window=(btn4))
    canvas.create_window(-100, -100, window=(btn5))
    canvas.create_window(-100, -100, window=(btn6))
    canvas.create_window(-100, -100, window=(btn7))
    # canvas.create_window(-100, -100, window=(absen))

# GUI / ------------------------------------------------------------------------------------
root = tk.Tk()
root.title('Smart Absensi')
# size screen
window_width = 600
window_height = 500

# get the screen dimension
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# find the center point
center_x = int(screen_width/2 - window_width / 2)
center_y = int(screen_height/2 - window_height / 2)

# set the position of the window to the center of the screen
root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

#ukuran layar tidak dapat diubah
root.resizable(False, False)

#layar menjadi tranparan
#root.attributes('-alpha', 0.5)

#urutan penumpukan jendela
#root.attributes('-topmost', 1)


# mengatur canvas (window tkinter)
canvas = tk.Canvas(root, width=600, height=500)
canvas.grid(columnspan=1, rowspan=1)
canvas.configure(bg="blue")
# judul
judul = tk.Label(root, text="Smart Absensi", font=("Roboto",25),bg="lightblue", fg="black", width=100)
canvas.create_window(300, 25, window=judul)

#HOME-------------------------------------
# tombol untuk absenMasuk
text = tk.StringVar()
btn = tk.Button(root, textvariable=text, font=("Roboto",14), bg="#20bebe", fg="white", height=4, width=18, command=absenMasuk)
text.set("Absensi Masuk")
canvas.create_window(190, 200, window=btn)

# tombol untuk absenKeluar
text1 = tk.StringVar()
btn1 = tk.Button(root, textvariable=text1, font=("Roboto",14), bg="#20bebe", fg="white", height=4, width=18, command=absenKeluar)
text1.set("Absensi Keluar")
canvas.create_window(410, 200, window=btn1)

# tombol untuk rekam data wajah
text2 = tk.StringVar()
btn2 = tk.Button(root, textvariable=text2, font=("Roboto",14), bg="#20bebe", fg="white", height=4, width=18, command=laporan)
text2.set("Laporan Absensi")
canvas.create_window(190, 330, window=btn2)

# tombol untuk rekam data wajah
text3 = tk.StringVar()
btn3 = tk.Button(root, textvariable=text3, font=("Roboto",14), bg="#20bebe", fg="white", height=4, width=18, command=daftar)
text3.set("Pendaftaran")
canvas.create_window(410, 330, window=btn3)


#PENDAFTARAN--------------------------------------------
label1 = tk.Label(root, text="Nama Pegawai", font=("Roboto",14), fg="white", bg="blue")
canvas.create_window(-100,-100, window=label1)
# for entry data nama
entry1 = tk.Entry (root, font="Roboto")
canvas.create_window(-100, -100, height=25, width=411, window=entry1)
text4 = tk.StringVar()
btn4 = tk.Button(root, textvariable=text4, font=("Roboto",14), bg="#20bebe", fg="white", height=2, width=15, command=ambilGambar)
text4.set("Ambil Gambar")
canvas.create_window(-100, -100, window=btn4)
text6 = tk.StringVar()
btn6 = tk.Button(root, textvariable=text6, font=("Roboto",12), bg="#20bebe", fg="white", height=1, width=20, command=pegawai)
text6.set("Lihat Data Pegawai")
text7 = tk.StringVar()
btn7 = tk.Button(root, textvariable=text7, font=("Roboto",12), bg="#20bebe", fg="white", height=1, width=20, command=trainingWajah)
text7.set("Training Data")
canvas.create_window(-100, -100, window=btn6)
text5 = tk.StringVar()
btn5 = tk.Button(root, textvariable=text5, font=("Roboto",10), bg="blue", fg="white", height=1, width=8, command=home)
text5.set("Kembali")
canvas.create_window(-100, -100, window=btn5)

#absensi


home()

root.mainloop()