import csv
import os
from datetime import datetime
import time
import openpyxl
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl import load_workbook

# global variable
global bulan, hasil
now = datetime.now()
bulan = now.strftime('%m-%Y')
hari = now.strftime('%d-%m-%Y')
wb = Workbook()
ws = wb.active


def cek_data():
    path = 'rekapData'
    tgl = str(now.strftime('%m%Y'))
    imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
    lsdata = []
    for imagePath in imagePaths:
        a = str(os.path.split(imagePath)[1].split("-")[0])
        b = str(os.path.split(imagePath)[1].split("-")[1])
        c = str(a)+str(b)
        lsdata.append(c)
    global hasil
    hasil = "Tidak ada"
    for row in lsdata:
        row = str(row)
        if str(tgl) == str(row):
            hasil = "ada"
            break
        else:
            hasil = "Tidak ada"
    return hasil

bln = now.strftime('%m')

if bln[0]=="0":
	bln = bln[1]
	pass

def isLeapYear(year):
    if year > 1582:
        return year % 4 == 0 and year % 100 != 0 or year % 400 == 0
    return year % 4 == 0

def daysInMonth(year, month):
    if(year == 1582 and month == 10):
        return 21
    elif month == 2: 
        return 28 + isLeapYear(year) 
        return(31 - (((month - 1) % 7) % 2))

def cek_tanggal():
	jml_bln = (daysInMonth(2023, int(bln)))
	i=1
	header = ["Nama"]
	while i<=jml_bln:
		if i<10:
			n = "0"+str(i)
			header.append(n+"-"+bulan)
		else:
			header.append(str(i)+"-"+bulan)
		print(header)
		i+=1
	return header


def buatFile():
	dt = cek_data()
	header = []
	print(str(dt))

	if str(dt) != "ada":
		ws.title = 'Rekap Masuk'
		workaktif = wb['Rekap Masuk']
		now = datetime.now()

		#header = ["Nama",1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31]
		header = cek_tanggal()
		for row in range(0,len(header)):
			ws.cell(row=1, column=row+1).value= header[row]

		nama = []
		names=[]
		data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
		sheet = data.active
		i=1
		for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
			nama.append(value)
			i+=1
			i=0
		for npegawai in nama:
			if i!=0:
				ws.cell(row=i+1, column=1).value= npegawai[1]
			i+=1
		wb.save("rekapData/"+bulan+'-masuk.xlsx')

		#reksap keluar
		ws.title = 'Rekap Keluar'
		workaktif = wb['Rekap Keluar']
		now = datetime.now()
		#header = ["Nama",1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31]
		header = cek_tanggal()
		for row in range(0,len(header)):
			ws.cell(row=1, column=row+1).value= header[row]

		nama = []
		names=[]
		data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
		sheet = data.active
		i=1
		for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
			nama.append(value)
			i+=1
			i=0
		for npegawai in nama:
			if i!=0:
				ws.cell(row=i+1, column=1).value= npegawai[1]
			i+=1
		wb.save("rekapData/"+bulan+'-keluar.xlsx')


def rekapExcel(nama, waktu):
	buatFile()
	if waktu == "Masuk":
		tgl = str(now.strftime('%d'))
		data = load_workbook(filename="rekapData/"+bulan+"-masuk.xlsx")
		sheet = data.active
		baris = 0
		for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
			baris +=1
			for row in range(0,len(value)):
				ws.cell(row=baris, column=row+1).value= value[row]

		baris = 0
		for value in sheet.iter_rows(values_only=True):
			baris +=1
			for row in range(0,len(value)):
				if value[row] == nama :
					ws.cell(row=baris, column=1+int(tgl)).value= "Hadir"
					#"D:/fotoabsensi/"+bulan+"/"+id+"/"+dtString+ ".jpg"
		wb.save("rekapData/"+bulan+'-masuk.xlsx')
	elif waktu == "Keluar":
		tgl = str(now.strftime('%d'))
		data = load_workbook(filename="rekapData/"+bulan+"-keluar.xlsx")
		sheet = data.active
		baris = 0
		for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
			baris +=1
			for row in range(0,len(value)):
				ws.cell(row=baris, column=row+1).value= value[row]

		baris = 0
		for value in sheet.iter_rows(values_only=True):
			baris +=1
			for row in range(0,len(value)):
				if value[row] == nama :
					ws.cell(row=baris, column=1+int(tgl)).value= "Hadir"
					#"D:/fotoabsensi/"+bulan+"/"+id+"/"+dtString+ ".jpg"
		wb.save("rekapData/"+bulan+'-keluar.xlsx')

