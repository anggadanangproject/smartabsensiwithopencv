import cv2, os, numpy as np
import tkinter as tk
from tkinter import *
import csv
from PIL import ImageTk, Image
from tkinter import ttk
from tkinter.messagebox import showinfo
from datetime import datetime
import time
from random import randint
import simpanAbsensi as rekap
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

global dt_abs, now, dtString
now = datetime.now()
dtString = now.strftime('%d-%m-%Y')
bulan = now.strftime('%m-%Y')
bil = 0

def random():
    bil = randint(3, 10)
    return bil

def cek_data():
    path = 'dataabsensi'
    tgl = str(now.strftime('%d%m'))
    imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
    lsdata = []
    for imagePath in imagePaths:
        a = str(os.path.split(imagePath)[1].split("-")[0])
        b = str(os.path.split(imagePath)[1].split("-")[1])
        c = str(a)+str(b)
        lsdata.append(c)

    global hasil
    hasil = "Tidak ada"
    for row in lsdata:
        #print(tgl+"/"+row)
        if str(tgl) == str(row):

            hasil = "ada"
            break
        else:
            hasil = "Tidak ada"
    #print(hasil)
    return hasil

def buat_file():
    now = datetime.now()
    dtS = now.strftime('%d-%m-%Y')
    dt = cek_data()
    nama = []
    pegawai=[]
    data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
    sheet = data.active
    i=1
    for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
        nama.append(value)
        i+=1
    i=0
    for npegawai in nama:
        if i!=0:
            pegawai.append(npegawai[1])
        i+=1
    if str(dt) != "ada":
        ws.title = 'Rekap absensi'
        workaktif = wb['Rekap absensi']
        now = datetime.now()
        lst_data = ["Nama","Jam Masuk","Jam Keluar"]
        i=1
        for header in lst_data:
            ws.cell(row=1, column=i).value= header
            i+=1
        i=1
        for namaPegawai in pegawai:
            i+=1
            ws.cell(row=i, column=1).value= namaPegawai

        wb.save("dataabsensi/"+dtS+'.xlsx')

def file_absensi():
    buat_file()

def markAttendance(nama, waktu):
    #print(cek_data())
    now = datetime.now()
    dtS = now.strftime('%d-%m-%Y')
    file_absensi()
    data = load_workbook(filename="dataabsensi/" + dtS + ".xlsx")
    sheet = data.active
    baris = 0
    for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
        baris +=1
        for row in range(0,len(value)):
            ws.cell(row=baris, column=row+1).value= value[row]
    baris = 0
    if waktu == "Masuk":
        for value in sheet.iter_rows(values_only=True):
            baris +=1
            for row in range(0,len(value)):
                if value[row] == nama :
                    ws.cell(row=baris, column=2).value= str(now.strftime('%H:%M:%S'))


    else:
        for value in sheet.iter_rows(values_only=True):
            baris +=1
            for row in range(0,len(value)):
                if value[row] == nama :
                    ws.cell(row=baris, column=3).value= now.strftime('%H:%M:%S')
    wb.save("dataabsensi/"+dtS+'.xlsx')


def folderFoto():
    path = 'D:/fotoabsensi/'
    tgl = str(now.strftime('%m%Y'))
    imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
    lsdata = []
    for imagePath in imagePaths:
        a = str(os.path.split(imagePath)[1].split("-")[0])
        b = str(os.path.split(imagePath)[1].split("-")[1])
        c = str(a)+str(b)
        lsdata.append(c)

    global hasil
    hasil = "Tidak ada"
    for row in lsdata:
        #print(tgl+"/"+row)
        if str(tgl) == str(row):

            hasil = "ada"
            break
        else:
            hasil = "Tidak ada"
    print(hasil)
    if hasil!="ada":
        os.mkdir("D:/fotoabsensi/"+bulan)
        nama = []
        names=[]
        data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
        sheet = data.active
        i=1
        for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
            nama.append(value)
            i+=1
            i=0
        for npegawai in nama:
            if i!=0:
                os.mkdir("D:/fotoabsensi/"+bulan+"/"+npegawai[1])
            i+=1

def absensiWajah(waktu):
    folderFoto()
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read('latihwajah/trainer.yml')
    cascadePath = "haarcascade_frontalface_default.xml"
    eye_cascade = cv2.CascadeClassifier('haarcascade_eye_tree_eyeglasses.xml')
    faceCascade = cv2.CascadeClassifier(cascadePath);

    first_read = True
    font = cv2.FONT_HERSHEY_SIMPLEX
    id = 0
    #yourname = entry1.get()

    # names related to ids: example ==> Marcelo: id=1,  etc
    nama = []
    names=[]
    data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
    sheet = data.active
    i=1
    for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
        nama.append(value)
        i+=1
    i=0
    for npegawai in nama:
        if i!=0:
            names.append(npegawai[1])
        i+=1
    print(names)
    # Initialize and start realtime video capture
    cam = cv2.VideoCapture(1)
    cam.set(3, 640)  # set video widht
    cam.set(4, 480)  # set video height

    # Define min window size to be recognized as a face
    minW = 0.1 * cam.get(3)
    minH = 0.1 * cam.get(4)
    deteksi = 0
    muncul=0
    absen = 0
    mata_terbuka = 0
    nama=""
    detik = random()
    lop = 0
    jeda = 0
    while True:

        ret, img = cam.read()
        img = cv2.flip(img, 1)  # Flip vertically

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale(
            gray,
            scaleFactor=1.2,
            minNeighbors=5,
            minSize=(int(minW), int(minH)),
        )
        
        if(len(faces)>0):
            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
                id, confidence = recognizer.predict(gray[y:y + h, x:x + w])
                #roi_face is face which is input to eye classifier
                roi_face = gray[y:y+h,x:x+w]
                roi_face_clr = img[y:y+h,x:x+w]
                eyes = eye_cascade.detectMultiScale(roi_face,1.3,5,minSize=(50,50))
                id = names[id-1]
                if(len(eyes)>=2):
                    lop+=1

                    if lop == 8:
                        detik = detik - 1
                        lop = 0
                        pass

                    if detik <= 0:
                        jeda+=1
                        mata_terbuka = 1
                        info =  "Kedipkan mata sekarang!"
                        if jeda == 10:
                            detik = random()
                            jeda = 0
                    if detik > 0:
                        mata_terbuka = 0
                        info = "Tunggu dalam "+str(detik)
                    
                    cv2.putText(img, info, (70,40), cv2.FONT_HERSHEY_PLAIN, 2,(0,255,0),2)

                else:
                    #To ensure if the eyes are present before starting
                    if (deteksi == 1 and mata_terbuka == 1):
                        absen = 1
                        cv2.putText(img, "Sudah absen!", (70,40), cv2.FONT_HERSHEY_PLAIN, 2,(0,0,255),2)

                    else:
                        mata_terbuka=0
                        absen=0
                        #cv2.putText(img, "Gagal absensi!", (70,40), cv2.FONT_HERSHEY_PLAIN, 2,(0,0,255),2)
                
                # Check if confidence is less them 100 ==> "0" is perfect match
                if (absen==1 and mata_terbuka ==1 and deteksi==1):
                    mata_terbuka = 0
                    absen = 0
                    nama = str(id)
                    print(str(id) + ", Kamu sudah absen!")
                    #selesai3(str(id))
                    markAttendance(str(id), waktu)
                    rekap.rekapExcel(str(id), waktu)
                    if waktu=="Masuk":
                        cv2.imwrite("D:/fotoabsensi/"+bulan+"/"+id+"/"+dtString+ "-masuk.jpg", img)
                    else:
                        cv2.imwrite("D:/fotoabsensi/"+bulan+"/"+id+"/"+dtString+ "-keluar.jpg", img)
                    break
                if (confidence <= 85):
                    deteksi = 1
                    confidence = "  {0}%".format(round(100 - confidence))
                else:
                    id = "tidak dikenali"
                    deteksi = 0
                    mata_terbuka=0
                    absen=0
                    confidence = "  {0}%".format(round(100 - confidence))
                cv2.putText(img, str(id), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
            cv2.putText(img, "Sudah absen: "+nama, (10,450), cv2.FONT_HERSHEY_PLAIN, 1,(255,0,255),1)
        else:
            deteksi=0
            mata_terbuka=0
            absen=0
            cv2.putText(img, "Wajah tidak terdeksi!", (20,20), cv2.FONT_HERSHEY_PLAIN, 1,(0,0,255),1)
            #info.config(text="Wajah tidak terdeksi!")

        muncul+=1

        cv2.imshow('camera', img)
        # if muncul >=30:
        #     if str(id).lower() == str(yourname).lower():
        #         print(str(id)+", Kamu sudah absen!")
        #         markAttendance(yourname)
        #         break
        k = cv2.waitKey(10) & 0xff  # Press 'ESC' for exiting video
        if k == 27:
                break

    # Do a bit of cleanup
    #print("\n [INFO] Exiting Program and cleanup stuff")
    cam.release()
    cv2.destroyAllWindows()


def absensi(waktu):
	absensiWajah(waktu)


def data():
    file_absensi()
    now = datetime.now()
    dtString = now.strftime('%d-%m-%Y')

    window = tk.Tk()
    window_width = 600
    window_height = 500
    # get the screen dimension
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # find the center point
    center_x = int(screen_width/2 - window_width / 2)
    center_y = int(screen_height/2 - window_height / 2)

    # set the position of the window to the center of the screen
    window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    #ukuran layar tidak dapat diubah
    window.resizable(False, False)
    window.title("Daftar Absensi / " + dtString)
    #urutan penumpukan jendela
    window.attributes('-topmost', 1)
    scroll_bar = Scrollbar(window)
    canvas = tk.Canvas(window, width=600, height=500)
    canvas.grid(columnspan=1, rowspan=1)
    canvas.configure(bg="blue")


    now = datetime.now()
    dtS = now.strftime('%d-%m-%Y')

    mylist = Listbox(window, font=("Roboto",14), fg="white", bg="blue", width=54,height=21, yscrollcommand = scroll_bar.set )
    data = load_workbook(filename="dataabsensi/"+dtS+".xlsx")
    sheet = data.active
    for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
        mylist.insert(END,str(value))

    canvas.create_window(300, 248, window=mylist)
    window.mainloop()
    

