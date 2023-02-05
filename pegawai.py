import cv2, os, numpy as np
import tkinter as tk
from tkinter import *
import csv
from PIL import ImageTk, Image
from tkinter import ttk
from tkinter.messagebox import showinfo
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active


global dt_abs, now, dtString
now = datetime.now()
dtString = now.strftime('%d-%m-%Y')

def data():
    now = datetime.now()
    dtString = now.strftime('%d-%m-%Y')
    window = tk.Tk()
    window_width = 600
    window_height = 500
    # get the screen dimension
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # find the center point
    center_x = int(screen_width/2 - window_width / 2)
    center_y = int(screen_height/2 - window_height / 2)

    # set the position of the window to the center of the screen
    window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    #ukuran layar tidak dapat diubah
    window.resizable(False, False)
    window.title("Daftar Pegawai / " + dtString)
    #urutan penumpukan jendela
    window.attributes('-topmost', 1)
    scroll_bar = Scrollbar(window)
    canvas = tk.Canvas(window, width=600, height=500)
    canvas.grid(columnspan=1, rowspan=1)
    canvas.configure(bg="blue")
    now = datetime.now()
    dtS = now.strftime('%d-%m-%Y')
    nama_abs = ""
    mylist = Listbox(window, font=("Roboto",14), fg="white", bg="blue", width=54,height=21, yscrollcommand = scroll_bar.set )

    data = load_workbook(filename="data_pegawai/data_pegawai.xlsx")
    sheet = data.active

    for value in sheet.iter_rows(min_row=1,min_col=1,values_only=True):
        mylist.insert(END,str(value))

    canvas.create_window(300, 248, window=mylist)
    window.mainloop()